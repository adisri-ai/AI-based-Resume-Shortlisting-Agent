from queue import Queue
from typing import Set
import threading
from azure.storage.blob import BlobServiceClient, BlobSasPermissions, generate_blob_sas
from typing import Set , List , Optional, Tuple
import io
from openpyxl import Workbook, load_workbook
import logging
import json
import datetime 
import zipfile
import os

class BlobManager : 
    def initialize(*args):
        BlobManager.__blob_queue = Queue()
        BlobManager.__processed_blobs  : Set[str] = []
        BlobManager.__processing_blobs : Set[str] = []
        BlobManager.__queue_lock = threading.Lock()
        BlobManager.___processing_status = {
            "total":     0,   
            "processed": 0,  
        }
        BlobManager.__status_lock = threading.Lock()
    @staticmethod
    def increment_processed_count(self):
        with BlobManager.__status_lock:
            BlobManager.___processing_status["processed"] += 1 
    @staticmethod
    def is_processed_or_processing(blob_name)->bool:
        return blob_name in BlobManager.__processed_blobs or blob_name in BlobManager.__processing_blobs
    @staticmethod
    def getQueue():
        return BlobManager.__blob_queue
    @staticmethod
    def getlock():
        return BlobManager.__queue_lock
    @staticmethod
    def mark_processing(blob_name):
        BlobManager.__processing_blobs.add(blob_name)
    @staticmethod
    def mark_processed(blob_name) : 
        BlobManager.__processed_blobs.add(blob_name)
        BlobManager.__processing_blobs.discard(blob_name)
    @staticmethod
    def blob_watcher_thread():
        pass
    
    @staticmethod
    def init_results_workbook(
        blob_service_client: BlobServiceClient,
        results_container: str,
        skills: List[str],
        RESULTS_BLOB_NAME = "results"
    ) -> None:
        container_client = blob_service_client.get_container_client(results_container)
        blob_client      = container_client.get_blob_client(RESULTS_BLOB_NAME)
        workbook         = Workbook()
        worksheet        = workbook.active
        worksheet.title  = "Results"

        header = ["CvName"] + skills + ["TotalScore", "ResumeUrl"]
        worksheet.append(header)

        out_mem_file = io.BytesIO()
        workbook.save(out_mem_file)
        out_mem_file.seek(0)
        blob_client.upload_blob(out_mem_file.read(), overwrite=True)
        logging.info(
            "Initialized '%s' in container '%s' with new JD skill columns.",
            RESULTS_BLOB_NAME, results_container,
        )

    @staticmethod
    def save_active_jd_metadata(
        blob_service_client: BlobServiceClient,
        config_container: str,
        jd_filename: str,
        skills: List[str],
        ACTIVE_JD_BLOB_NAME
    ) -> None:
        container_client = blob_service_client.get_container_client(config_container)
        blob_client      = container_client.get_blob_client(ACTIVE_JD_BLOB_NAME)
        meta = {
            "jd_filename":    jd_filename,
            "skills":         skills,
            "updated_at_utc": datetime.utcnow().isoformat(),
            "daily_scans":    [],
        }

        data = json.dumps(meta, indent=2)
        blob_client.upload_blob(data, overwrite=True)
        logging.info(
            "Updated '%s' in container '%s' with active JD '%s'.",
            ACTIVE_JD_BLOB_NAME, config_container, jd_filename,
        )
    @staticmethod
    def get_storage_account_info_from_connection_string(conn_str: str):
        """
        Parse account name and key from a standard Storage connection string.
        Example:
        DefaultEndpointsProtocol=https;AccountName=xxx;AccountKey=yyy;EndpointSuffix=core.windows.net
        """
        parts = dict(
            item.split("=", 1)
            for item in conn_str.split(";")
            if "=" in item
        )
        account_name = parts.get("AccountName")
        account_key  = parts.get("AccountKey")
        if not account_name or not account_key:
            raise ValueError("Could not parse AccountName/AccountKey from DATA_STORAGE_CONNECTION")
        return account_name, account_key
    @staticmethod
    def load_active_jd_metadata(
        blob_service_client: BlobServiceClient,
        config_container: str,
        ACTIVE_JD_BLOB_NAME = "status"
    ) -> Optional[dict]:
        container_client = blob_service_client.get_container_client(config_container)
        blob_client      = container_client.get_blob_client(ACTIVE_JD_BLOB_NAME)
        try:
            download = blob_client.download_blob()
            data     = download.readall().decode("utf-8")
            meta     = json.loads(data)
            return meta
        except Exception:
            return None

    @staticmethod
    def update_active_jd_scan_log(
            blob_service_client: BlobServiceClient,
            config_container: str,
            cv_filename: str,
            ACTIVE_JD_BLOB_NAME = "status"
        ) -> None:
            meta = BlobManager.load_active_jd_metadata(blob_service_client, config_container, ACTIVE_JD_BLOB_NAME)
            if not meta:
                return

            today_str   = datetime.utcnow().date().isoformat()
            daily_scans = meta.get("daily_scans", [])

            updated = False
            for entry in daily_scans:
                if entry.get("date") == today_str:
                    cvs = entry.get("cvs", [])
                    if cv_filename not in cvs:
                        cvs.append(cv_filename)
                        entry["cvs"]   = cvs
                        entry["count"] = len(cvs)
                    updated = True
                    break

            if not updated:
                daily_scans.append({
                    "date":  today_str,
                    "count": 1,
                    "cvs":   [cv_filename],
                })

            meta["daily_scans"] = daily_scans

            container_client = blob_service_client.get_container_client(config_container)
            blob_client      = container_client.get_blob_client(ACTIVE_JD_BLOB_NAME)
            blob_client.upload_blob(json.dumps(meta, indent=2), overwrite=True)
            logging.info(
                "Updated daily_scans in '%s' with CV '%s' for date %s.",
                ACTIVE_JD_BLOB_NAME, cv_filename, today_str,
            )
    
    @staticmethod
    def read_skills_from_results_header(
        blob_service_client: BlobServiceClient,
        results_container: str,
        RESULTS_BLOB_NAME = "results"
    ) -> Optional[List[str]]:
        container_client = blob_service_client.get_container_client(results_container)
        blob_client      = container_client.get_blob_client(RESULTS_BLOB_NAME)
        try:
            download_stream = blob_client.download_blob()
            data            = download_stream.readall()
            in_mem_file     = io.BytesIO(data)
            workbook        = load_workbook(in_mem_file)
            worksheet       = workbook.active
            header_row      = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            header          = list(header_row)
            if len(header) < 3:
                logging.error(
                    "results.xlsx header has fewer than 3 columns; cannot read skills."
                )
                return None
            try:
                total_idx = header.index("TotalScore")
                skills    = header[1:total_idx]
            except ValueError:
                skills    = header[1:-1]
            skills = [s for s in skills if s]
            return skills
        except Exception as e:
            logging.error("Failed to read skills from results.xlsx header: %s", e)
            return None
    @staticmethod
    def _build_resume_url(cv_filename: str) -> str:
        try:
            conn_str           = os.environ.get("DATA_STORAGE_CONNECTION", "")
            account_name, _    = BlobManager.get_storage_account_info_from_connection_string(conn_str)
            incoming_container = os.environ.get("BLOB_INCOMING_CONTAINER", "incoming")
            return (
                f"https://{account_name}.blob.core.windows.net"
                f"/{incoming_container}/{cv_filename}"
            )
        except Exception:
            return "" 
    @staticmethod
    def cv_already_scored(
        blob_service_client: BlobServiceClient,
        results_container: str,
        cv_filename: str,
        RESULTS_BLOB_NAME = "results"
        ) -> bool:
            container_client = blob_service_client.get_container_client(results_container)
            blob_client      = container_client.get_blob_client(RESULTS_BLOB_NAME)

            try:
                download_stream = blob_client.download_blob()
                data            = download_stream.readall()
                in_mem_file     = io.BytesIO(data)
                workbook        = load_workbook(in_mem_file)
                worksheet       = workbook.active
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    existing_name = row[0]
                    if existing_name and str(existing_name) == cv_filename:
                        return True
            except Exception:
                return False

            return False
    
    @staticmethod
    def append_cv_scores_to_results(
        blob_service_client: BlobServiceClient,
        results_container: str,
        cv_filename: str,
        scores: List[float],
        total_score: float,
        resume_url: str = "", 
        RESULTS_BLOB_NAME = "results"
    ) -> None:
        container_client = blob_service_client.get_container_client(results_container)
        blob_client      = container_client.get_blob_client(RESULTS_BLOB_NAME)

        max_retries = 5
        for attempt in range(max_retries):
            try:
                downloader = blob_client.download_blob()
                etag       = downloader.properties.etag
                data       = downloader.readall()
            except Exception as e:
                logging.error(
                    "Failed to download '%s' for appending CV '%s': %s",
                    RESULTS_BLOB_NAME, cv_filename, e,
                )
                return

            in_mem_file = io.BytesIO(data)
            workbook    = load_workbook(in_mem_file)
            worksheet   = workbook.active

            already = False
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                existing_name = row[0]
                if existing_name and str(existing_name) == cv_filename:
                    already = True
                    break

            if already:
                logging.info(
                    "CV '%s' is already present in results.xlsx; skipping append.",
                    cv_filename,
                )
                return
            row_data = [cv_filename] + list(scores) + [total_score, resume_url]
            worksheet.append(row_data)

            out_mem_file = io.BytesIO()
            workbook.save(out_mem_file)
            out_mem_file.seek(0)
            new_data = out_mem_file.read()

            try:
                blob_client.upload_blob(
                    new_data,
                    overwrite=True,
                    if_match=etag,
                )
                logging.info(
                    "Appended CV '%s' scores to results.xlsx (TotalScore=%.2f).",
                    cv_filename, total_score,
                )
                return
            except Exception as e:
                logging.warning(
                    "Concurrency conflict appending '%s' to results.xlsx (attempt %d/%d): %s",
                    cv_filename, attempt + 1, max_retries, e,
                )

        logging.error(
            "Failed to append CV '%s' to results.xlsx after %d retries.",
            cv_filename, max_retries,
        )
    
    @staticmethod
    def handle_zip_upload(
        blob_service_client: BlobServiceClient,
        incoming_container: str,
        blob_filename: str,
        blob_bytes: bytes,
    ) -> None:
        logging.info("Handling zip upload for '%s'.", blob_filename)
        extracted_files = []
        pdf_count       = 0

        try:
            with zipfile.ZipFile(io.BytesIO(blob_bytes)) as zf:
                for info in zf.infolist():
                    if info.is_dir():
                        continue

                    inner_name = info.filename
                    logging.info(
                        "Extracting '%s' from zip '%s' into container '%s'.",
                        inner_name, blob_filename, incoming_container,
                    )

                    file_bytes = zf.read(info)
                    if not file_bytes:
                        continue

                    container_client   = blob_service_client.get_container_client(incoming_container)
                    inner_blob_client  = container_client.get_blob_client(inner_name)
                    inner_blob_client.upload_blob(file_bytes, overwrite=True)
                    extracted_files.append(inner_name)

                    if inner_name.lower().endswith(".pdf"):
                        pdf_count += 1

        except Exception as e:
            logging.error("Failed to extract zip '%s': %s", blob_filename, e)
            return
        with BlobManager.__status_lock:
            BlobManager.__processing_status["total"]    = pdf_count
            BlobManager.__processing_status["processed"] = 0

        logging.info(
            "Finished handling zip upload '%s'; extracted %d files (%d PDFs).",
            blob_filename, len(extracted_files), pdf_count,
        )
    
    def get_scanned_and_pending_sets(
        blob_service_client: BlobServiceClient,
        status_container: str,
        results_container: str,
        config_container: str,
        RESULTS_BLOB_NAME = "results",
        OTHERS_STATUS_BLOB_NAME= "others",
        PENDING_STATUS_BLOB_NAME = "status"
    ) -> Tuple[Set[str], Set[str]]:
        scanned: Set[str] = set()
        pending: Set[str] = set()

        try:
            container_client = blob_service_client.get_container_client(results_container)
            blob_client      = container_client.get_blob_client(RESULTS_BLOB_NAME)
            download_stream  = blob_client.download_blob()
            data             = download_stream.readall()
            in_mem_file      = io.BytesIO(data)
            workbook         = load_workbook(in_mem_file)
            worksheet        = workbook.active
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                cv_name = row[0]
                if cv_name:
                    scanned.add(str(cv_name))
        except Exception:
            pass
        others = BlobManager.read_filenames_from_status_list(
            blob_service_client, status_container, OTHERS_STATUS_BLOB_NAME
        )
        scanned.update(others)

        jd_meta = BlobManager.load_active_jd_metadata(blob_service_client, config_container)
        if jd_meta:
            old_jd = jd_meta.get("jd_filename")
            if old_jd:
                scanned.add(str(old_jd))

        pending = BlobManager.read_filenames_from_status_list(
            blob_service_client, status_container, PENDING_STATUS_BLOB_NAME
        )

        return scanned, pending
    
    @staticmethod 
    def append_filename_to_status_list(
            blob_service_client: BlobServiceClient,
            status_container: str,
            list_blob_name: str,
            filename: str,
        ) -> None:
            container_client =  blob_service_client.get_container_client(status_container)
            blob_client      = container_client.get_blob_client(list_blob_name)
            line = f"{filename}\n"
            try:
                try:
                    existing = blob_client.download_blob().readall()
                except Exception:
                    existing = b""
                new_content = existing + line.encode("utf-8")
                blob_client.upload_blob(new_content, overwrite=True)
                logging.info(
                    "Appended '%s' to status list '%s/%s'.",
                    filename, status_container, list_blob_name,
                )
            except Exception as e:
                logging.error(
                    "Failed to append '%s' to status list '%s': %s",
                    filename, list_blob_name, e,
                )

    @staticmethod
    def read_filenames_from_status_list(
        blob_service_client: BlobServiceClient,
        status_container: str,
        list_blob_name: str,
    ) -> Set[str]:
        container_client = blob_service_client.get_container_client(status_container)
        blob_client      = container_client.get_blob_client(list_blob_name)
        names: Set[str]  = set()
        try:
            data = blob_client.download_blob().readall().decode("utf-8")
            for line in data.splitlines():
                name = line.strip()
                if name:
                    names.add(name)
        except Exception:
            pass
        return names
    
    @staticmethod
    def handle_jd(
    blob_service_client: BlobServiceClient,
    blob_filename: str,
    full_text: str,
    incoming_container: str,
    status_container: str,
    results_container: str,
    config_container: str,
    skills : List[str]
    ) -> None:
        logging.info("Handling JD '%s'.", blob_filename)
        BlobManager.__processing_status["total"] = 0
        BlobManager.__processing_status["processed"] = 0
        bsc      = blob_service_client
        incoming = os.environ.get("BLOB_INCOMING_CONTAINER", "incoming")
        container_client = bsc.get_container_client(
            incoming
        )
        for blob in container_client.list_blobs():
            if(blob.name == blob_filename):
                continue
            container_client.delete_blob(blob.name)
            with BlobManager.__queue_lock:
                BlobManager.__processed_blobs.discard(blob.name)
                BlobManager.__processing_blobs.discard(blob.name)
        scanned, pending = BlobManager.get_scanned_and_pending_sets(
            blob_service_client, status_container, results_container, config_container
        )

        inc_client = blob_service_client.get_container_client(incoming_container)
        for blob in inc_client.list_blobs():
            base_name = os.path.basename(blob.name)
            if base_name == blob_filename:
                continue
            if base_name in pending:
                continue
            if base_name in scanned:
                logging.info(
                    "Deleting previously scanned blob '%s' (base '%s') from '%s'.",
                    blob.name, base_name, incoming_container,
                )
                inc_client.delete_blob(blob.name)

        BlobManager.init_results_workbook(blob_service_client, results_container, skills)
        BlobManager.save_active_jd_metadata(blob_service_client, config_container, blob_filename, skills)

        logging.info(
            "JD '%s' processed. Skills: %s. results.xlsx re-initialized.",
            blob_filename, ", ".join(skills),
        )

    @staticmethod
    def handle_cv(
        blob_service_client: BlobServiceClient,
        blob_filename: str,
        full_text: str,
        status_container: str,
        results_container: str,
        config_container: str,
        PENDING_STATUS_BLOB_NAME = "status",
        scores = 0,
        total_score = 0
    ) -> None:
        logging.info("Handling CV '%s'.", blob_filename)

        if BlobManager.cv_already_scored(blob_service_client, results_container, blob_filename):
            logging.info(
                "CV '%s' is already scored in results.xlsx; skipping re-processing.",
                blob_filename,
            )
            return

        jd_meta = BlobManager.load_active_jd_metadata(blob_service_client, config_container)
        jd_name = None
        if jd_meta:
            jd_name = jd_meta.get("jd_filename")
        if not jd_meta or not jd_name:
            logging.warning(
                "No valid active JD found; CV '%s' will be marked as pending.", blob_filename
            )
            BlobManager.append_filename_to_status_list(
                blob_service_client, status_container, PENDING_STATUS_BLOB_NAME, blob_filename
            )
            return

        skills = BlobManager.read_skills_from_results_header(blob_service_client, results_container)
        if not skills or len(skills) != 10:
            logging.error(
                "Could not read 10 skills from results.xlsx header; found: %s. "
                "CV '%s' will not be scored.",
                skills, blob_filename,
            )
            return

        resume_url = BlobManager._build_resume_url(blob_filename)

        BlobManager.append_cv_scores_to_results(
            blob_service_client, results_container, blob_filename, scores, total_score,
            resume_url=resume_url,
        )

        BlobManager.update_active_jd_scan_log(blob_service_client, config_container, blob_filename)


        logging.info(
            "CV '%s' scored against active JD '%s'. TotalScore=%.2f.",
            blob_filename, jd_name, total_score,
        )
    
    @staticmethod
    def handle_other(
        blob_service_client: BlobServiceClient,
        blob_filename: str,
        status_container: str,
        results_container: str,
    ) -> None:
        logging.info(
            "File '%s' was classified as OTHER. Adding zero row to results.xlsx.",
            blob_filename,
        )
        OTHERS_STATUS_BLOB_NAME  = os.environ.get("OTHERS_STATUS_BLOB_NAME" , "other")
        BlobManager.append_filename_to_status_list(
            blob_service_client,
            status_container,
            OTHERS_STATUS_BLOB_NAME,
            blob_filename,
        )

        skills = BlobManager.read_skills_from_results_header(
            blob_service_client,
            results_container,
        )

        if skills:
            zero_scores = [0.0] * len(skills)
            resume_url  = BlobManager._build_resume_url(blob_filename)
            BlobManager.append_cv_scores_to_results(
                blob_service_client,
                results_container,
                blob_filename,
                zero_scores,
                0.0,
                resume_url=resume_url,
            )