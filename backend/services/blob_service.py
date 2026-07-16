from managers.blob_manager import BlobManager
import os
import logging
from azure.storage.blob import BlobServiceClient, BlobSasPermissions, generate_blob_sas
from typing import Set , List , Optional
import io
from openpyxl import Workbook , load_workbook
class BlobService : 
    @staticmethod
    def get_storage_account_info_from_connection_string(conn_str: str):
        return BlobManager.get_storage_account_info_from_connection_string(conn_str)
    @staticmethod
    def getQueue():
        return BlobManager.getQueue()
    @staticmethod
    def getlock():
        return BlobManager.getlock()
    @staticmethod
    def get_blob_service_client() -> BlobServiceClient:
        conn_str = os.environ["DATA_STORAGE_CONNECTION"]
        return BlobServiceClient.from_connection_string(conn_str)
    
    @staticmethod
    def get_container_client():
        pass
    @staticmethod
    def mark_processing(blob_name):
        BlobManager.mark_processing(blob_name)
    @staticmethod
    def mark_processed(blob_name):
        BlobManager.mark_processed(blob_name)
    @staticmethod
    def increment_processed_count():
        BlobManager.increment_processed_count()
    @staticmethod
    def is_processed_or_processing(blob_name) -> bool:
        return BlobManager.is_processed_or_processing(blob_name)
    @staticmethod
    def initialize():
        BlobManager.initialize()
        incoming = os.environ.get("BLOB_INCOMING_CONTAINER", "incoming")
        blob_service = BlobService.get_blob_service_client()
        incoming_container = blob_service.get_container_client()[incoming]
        for blob in incoming_container.list_blobs():
            BlobService.mark_processed(blob.name)
            incoming_container.delete_blob(blob.name)

    @staticmethod
    def worker_thread():
        return BlobManager.blob_worker_thread
    
    @staticmethod
    def watcher_thread():
        return BlobManager.blob_watcher_thread
    
    @staticmethod
    def cleanup_incoming_container():
        try:
            blob_service_client = BlobService.get_blob_service_client()

            incoming_container = os.environ.get(
                "BLOB_INCOMING_CONTAINER",
                "incoming",
            )

            container_client = blob_service_client.get_container_client(
                incoming_container
            )

            deleted_count = 0

            for blob in container_client.list_blobs():
                container_client.delete_blob(blob.name)
                deleted_count += 1

            logging.info(
                "Startup cleanup completed. Deleted %d blobs from '%s'.",
                deleted_count,
                incoming_container,
            )

        except Exception as e:
            logging.error("Failed startup cleanup: %s", e)

    @staticmethod
    def get_storage_account_info_from_connection_string(conn_str: str):
        """
        Parse account name and key from a standard Storage connection string.
        Example:
        DefaultEndpointsProtocol=https;AccountName=xxx;AccountKey=yyy;EndpointSuffix=core.windows.net
        """
        parts = dict(
            item.split("=", 1)
            for item in conn_str.split(";")
            if "=" in item
        )
        account_name = parts.get("AccountName")
        account_key  = parts.get("AccountKey")
        if not account_name or not account_key:
            raise ValueError("Could not parse AccountName/AccountKey from DATA_STORAGE_CONNECTION")
        return account_name, account_key
    @staticmethod
    def init_results_workbook(
        blob_service_client: BlobServiceClient,
        results_container: str,
        skills: List[str],
        RESULTS_BLOB_NAME = "results"
    ) -> None:
       BlobManager.init_results_workbook(blob_service_client , results_container , skills , RESULTS_BLOB_NAME)
    @staticmethod
    def save_active_jd_metadata(
        blob_service_client: BlobServiceClient,
        config_container: str,
        jd_filename: str,
        skills: List[str],
        ACTIVE_JD_BLOB_NAME = "status"
    ) -> None:
        BlobManager.save_active_jd_metadata(blob_service_client , config_container , jd_filename , skills , ACTIVE_JD_BLOB_NAME)
    @staticmethod
    def load_active_jd_metadata(
        blob_service_client: BlobServiceClient,
        config_container: str,
        ACTIVE_JD_BLOB_NAME = "status"
    ) -> Optional[dict]:
        return BlobManager.load_active_jd_metadata(blob_service_client , config_container , ACTIVE_JD_BLOB_NAME)
    @staticmethod
    def update_active_jd_scan_log(
            blob_service_client: BlobServiceClient,
            config_container: str,
            cv_filename: str,
            ACTIVE_JD_BLOB_NAME = "status"
        ) -> None:
        return BlobManager.update_active_jd_scan_log(blob_service_client, config_container,cv_filename,ACTIVE_JD_BLOB_NAME)
    
    @staticmethod
    def read_skills_from_results_header(
        blob_service_client: BlobServiceClient,
        results_container: str,
        RESULTS_BLOB_NAME = "results"
    ) -> Optional[List[str]]:
        return BlobManager.read_skills_from_results_header(blob_service_client , results_container, RESULTS_BLOB_NAME)
    
    @staticmethod
    def cv_already_scored(
    blob_service_client: BlobServiceClient,
    results_container: str,
    cv_filename: str,
    RESULTS_BLOB_NAME = "results"
    ) -> bool:
        return BlobManager.cv_already_scored(blob_service_client , results_container, cv_filename, RESULTS_BLOB_NAME)
    
    @staticmethod
    
    def append_cv_scores_to_results(
        blob_service_client: BlobServiceClient,
        results_container: str,
        cv_filename: str,
        scores: List[float],
        total_score: float,
        resume_url: str = "",   
    ) -> None:
        return BlobManager.append_cv_scores_to_results(blob_service_client, results_container, cv_filename,
                                                       scores, total_score, resume_url)
    
    @staticmethod
    def handle_zip_upload(
        blob_service_client: BlobServiceClient,
        incoming_container: str,
        blob_filename: str,
        blob_bytes: bytes,
    ) -> None:
        return BlobManager.handle_zip_upload(blob_service_client , incoming_container, blob_filename, blob_bytes)
    
    @staticmethod
    def handle_jd_upload(
        blob_service_client: BlobServiceClient,
        blob_filename: str,
        full_text: str,
        incoming_container: str,
        status_container: str,
        results_container: str,
        config_container: str,
        skills : List[str]
    ) -> None:
        BlobManager.handle_jd(blob_service_client , blob_filename , full_text , incoming_container, 
                              status_container, results_container, config_container, skills)
    
    @staticmethod  
    def handle_cv_upload(
        blob_service_client : BlobServiceClient,
        blob_filename: str,
        full_text: str,
        status_container: str,
        results_container: str,
        config_container: str,
        scores, total_score
    ):
        pending_blob = os.environ.get("PENDING_STATUS_BLOB" , "status")
        BlobManager.handle_cv(blob_service_client,blob_filename,
                              full_text,status_container,results_container,config_container,pending_blob,scores,total_score)
        
    @staticmethod
    def handle_other(
        blob_service_client: BlobServiceClient,
        blob_filename: str,
        status_container: str,
        results_container: str,
    ) -> None:
        BlobManager.handle_other(blob_service_client, blob_filename, status_container,results_container)    

    @staticmethod
    def get_results_json():
        blob_service_client = BlobService.get_blob_service_client()
        results_container   = os.environ.get("BLOB_RESULTS_CONTAINER", "results")
        RESULTS_BLOB_NAME = os.environ.get("RESULTS_BLOB_NAME" , "results")
        blob_client = (
            blob_service_client
            .get_container_client(results_container)
            .get_blob_client(RESULTS_BLOB_NAME)
        )

        rows   = []
        skills = []

        try:
            download_stream = blob_client.download_blob()
            data            = download_stream.readall()
            in_mem_file     = io.BytesIO(data)
            workbook        = load_workbook(in_mem_file)
            worksheet       = workbook.active

            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            header     = list(header_row)

            # Robustly find TotalScore and ResumeUrl positions
            try:
                total_idx      = header.index("TotalScore")
                resume_url_idx = header.index("ResumeUrl")
            except ValueError:
                total_idx      = len(header) - 1
                resume_url_idx = None

            skills = [h for h in header[1:total_idx] if h]

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                resume_url = ""
                if resume_url_idx is not None and len(row) > resume_url_idx:
                    resume_url = row[resume_url_idx] or ""
                rows.append({
                    "cvName":     row[0],
                    "scores":     list(row[1:total_idx]),
                    "totalScore": row[total_idx],
                    "resumeUrl":  resume_url,   # REQ 5
                })

        except Exception as e:
            logging.error("Failed to read results.xlsx: %s", e)

        return {"rows": rows, "skills": skills}